from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date

from app import models, schemas
from app.database import get_db
from app.dependencies import (
    get_current_admin,
    get_current_setter_or_admin,
    get_current_super_admin,
)

BOULDER_GRADES = [
    "V0",
    "V1",
    "V2",
    "V3",
    "V4",
    "V5",
    "V6",
    "V7",
    "V8",
    "V9",
    "V10",
    "V11",
    "V12",
]
TOP_ROPE_GRADES = [
    "5.5",
    "5.6",
    "5.7",
    "5.8",
    "5.9",
    "5.10a",
    "5.10b",
    "5.10c",
    "5.10d",
    "5.11a",
    "5.11b",
    "5.11c",
    "5.11d",
    "5.12a",
    "5.12b",
    "5.12c",
    "5.12d",
]
GRADES_BY_TYPE = {"boulder": BOULDER_GRADES, "top_rope": TOP_ROPE_GRADES}


def _validate_grade_for_zone(grade: str, zone: models.Zone) -> None:
    """Raise HTTPException if grade is invalid for the zone's route type."""
    valid_grades = GRADES_BY_TYPE.get(zone.route_type, [])
    if grade and valid_grades and grade not in valid_grades:
        raise HTTPException(
            status_code=400,
            detail=f"Grade '{grade}' is not valid for zone type '{zone.route_type}'. Valid grades: {valid_grades}",
        )


router = APIRouter(prefix="/admin", tags=["admin"])

# --- User Management ---


@router.get("/users", response_model=List[schemas.UserResponse])
def get_users(
    role: Optional[str] = None,
    is_banned: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    query = db.query(models.User)
    if role:
        query = query.filter(models.User.role == role)
    if is_banned is not None:
        query = query.filter(models.User.is_banned == is_banned)
    return query.all()


@router.patch("/users/{user_id}/ban", response_model=schemas.UserResponse)
def update_user_ban_status(
    user_id: int,
    ban_update: schemas.UserBanUpdate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if db_user.role == "super_admin":
        raise HTTPException(status_code=403, detail="Cannot ban super admins")
    if current_admin.role == "admin" and db_user.role == "admin":
        raise HTTPException(status_code=403, detail="Admins cannot ban other admins")
    if db_user.id == current_admin.id:
        raise HTTPException(status_code=403, detail="Cannot ban yourself")

    db_user.is_banned = ban_update.is_banned
    db.commit()
    db.refresh(db_user)
    return db_user


@router.patch("/users/{user_id}/role", response_model=schemas.UserResponse)
def update_user_role(
    user_id: int,
    role_update: schemas.UserRoleUpdate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")

    valid_roles = ["student", "setter", "admin", "super_admin"]
    if role_update.role not in valid_roles:
        raise HTTPException(status_code=400, detail="Invalid role")

    if current_admin.role == "admin":
        if role_update.role in ["admin", "super_admin"] or db_user.role in [
            "admin",
            "super_admin",
        ]:
            raise HTTPException(
                status_code=403, detail="Admins can only manage students and setters"
            )
    elif current_admin.role == "super_admin":
        if db_user.id == current_admin.id and role_update.role != "super_admin":
            raise HTTPException(status_code=403, detail="Cannot demote yourself")

    db_user.role = role_update.role
    db.commit()
    db.refresh(db_user)
    return db_user


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_super_admin: models.User = Depends(get_current_super_admin),
):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")

    if db_user.id == current_super_admin.id:
        raise HTTPException(status_code=403, detail="Cannot delete yourself")

    db.delete(db_user)
    db.commit()
    return None


# --- Zone Management ---


@router.post(
    "/zones", response_model=schemas.ZoneResponse, status_code=status.HTTP_201_CREATED
)
def create_zone(
    zone: schemas.ZoneCreate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_zone = models.Zone(
        name=zone.name, description=zone.description, route_type=zone.route_type
    )
    db.add(db_zone)
    db.commit()
    db.refresh(db_zone)
    return db_zone


# --- Color Management ---


@router.post(
    "/colors", response_model=schemas.ColorResponse, status_code=status.HTTP_201_CREATED
)
def create_color(
    color: schemas.ColorCreate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_color = models.Color(name=color.name, hex_value=color.hex_value)
    db.add(db_color)
    try:
        db.commit()
        db.refresh(db_color)
    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=400, detail="Color with this name might already exist"
        )
    return db_color


@router.delete("/colors/{color_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_color(
    color_id: int,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_color = db.query(models.Color).filter(models.Color.id == color_id).first()
    if not db_color:
        raise HTTPException(status_code=404, detail="Color not found")

    db.delete(db_color)
    db.commit()
    return None


# --- Setter Management ---


@router.get("/setters", response_model=List[schemas.SetterResponse])
def get_setters(
    name: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_setter_or_admin),
):
    query = db.query(models.Setter)
    if name:
        query = query.filter(models.Setter.name.ilike(f"%{name}%"))
    if is_active is not None:
        query = query.filter(models.Setter.is_active == is_active)
    return query.all()


@router.post(
    "/setters",
    response_model=schemas.SetterResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_setter(
    setter: schemas.SetterCreate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_setter = models.Setter(name=setter.name, is_active=setter.is_active)
    db.add(db_setter)
    db.commit()
    db.refresh(db_setter)
    return db_setter


@router.patch("/setters/{setter_id}", response_model=schemas.SetterResponse)
def update_setter(
    setter_id: int,
    setter_update: schemas.SetterUpdate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_setter = db.query(models.Setter).filter(models.Setter.id == setter_id).first()
    if not db_setter:
        raise HTTPException(status_code=404, detail="Setter not found")
    if setter_update.is_active is not None:
        db_setter.is_active = setter_update.is_active
    if setter_update.name is not None:
        db_setter.name = setter_update.name
    db.commit()
    db.refresh(db_setter)
    return db_setter


@router.delete("/setters/{setter_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_setter(
    setter_id: int,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_setter = db.query(models.Setter).filter(models.Setter.id == setter_id).first()
    if not db_setter:
        raise HTTPException(status_code=404, detail="Setter not found")

    db.delete(db_setter)
    db.commit()
    return None


# --- Route Management ---


@router.post(
    "/routes", response_model=schemas.RouteResponse, status_code=status.HTTP_201_CREATED
)
def create_route(
    route: schemas.RouteCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_setter_or_admin),
):
    zone = db.query(models.Zone).filter(models.Zone.id == route.zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")

    if route.setter_id is not None:
        setter = (
            db.query(models.Setter).filter(models.Setter.id == route.setter_id).first()
        )
        if not setter:
            raise HTTPException(status_code=404, detail="Setter not found")

    _validate_grade_for_zone(route.intended_grade, zone)

    db_route = models.Route(
        setter_id=route.setter_id,
        zone_id=route.zone_id,
        color=route.color,
        intended_grade=route.intended_grade,
        status=route.status,
    )
    if route.set_date is not None:
        db_route.set_date = route.set_date
    db.add(db_route)
    db.commit()
    db.refresh(db_route)
    return db_route


@router.patch("/routes/{route_id}", response_model=schemas.RouteResponse)
def update_route(
    route_id: int,
    route_update: schemas.RouteUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_setter_or_admin),
):
    db_route = db.query(models.Route).filter(models.Route.id == route_id).first()
    if not db_route:
        raise HTTPException(status_code=404, detail="Route not found")

    if route_update.zone_id is not None:
        zone = (
            db.query(models.Zone).filter(models.Zone.id == route_update.zone_id).first()
        )
        if not zone:
            raise HTTPException(status_code=404, detail="Zone not found")
        db_route.zone_id = route_update.zone_id

    if route_update.setter_id is not None:
        setter = (
            db.query(models.Setter)
            .filter(models.Setter.id == route_update.setter_id)
            .first()
        )
        if not setter:
            raise HTTPException(status_code=404, detail="Setter not found")
        db_route.setter_id = route_update.setter_id

    if route_update.color is not None:
        db_route.color = route_update.color
    if route_update.intended_grade is not None:
        zone = (
            db.query(models.Zone)
            .filter(models.Zone.id == (route_update.zone_id or db_route.zone_id))
            .first()
        )
        if zone:
            _validate_grade_for_zone(route_update.intended_grade, zone)
        db_route.intended_grade = route_update.intended_grade
    if route_update.status is not None:
        db_route.status = route_update.status
    if route_update.set_date is not None:
        db_route.set_date = route_update.set_date

    db.commit()
    db.refresh(db_route)
    return db_route


@router.patch("/routes/{route_id}/archive", response_model=schemas.RouteResponse)
def archive_route(
    route_id: int,
    archive_data: schemas.RouteArchive,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_setter_or_admin),
):
    db_route = db.query(models.Route).filter(models.Route.id == route_id).first()
    if not db_route:
        raise HTTPException(status_code=404, detail="Route not found")

    db_route.status = archive_data.status
    db.commit()
    db.refresh(db_route)
    return db_route


from openpyxl.worksheet.datavalidation import DataValidation


@router.get("/routes/bulk-template")
def get_bulk_route_template(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_super_admin),
):
    wb = openpyxl.Workbook()

    # Reference Sheet (Create this first so formulas can reference it)
    ws_ref = wb.active
    ws_ref.title = "Reference"

    header_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    header_font = Font(bold=True)

    ws_ref.append(["Zones (with type)", "Setters", "Colors", "All Grades"])
    for col in range(1, 4):
        cell = ws_ref.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        ws_ref.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    zones = db.query(models.Zone).all()
    setters = db.query(models.Setter).filter(models.Setter.is_active == True).all()
    colors = db.query(models.Color).all()
    all_grades = BOULDER_GRADES + TOP_ROPE_GRADES

    max_rows = max(len(zones), len(setters), len(colors), len(all_grades))

    for i in range(max_rows):
        row = []
        row.append(f"{zones[i].name}" if i < len(zones) else "")
        row.append(setters[i].name if i < len(setters) else "")
        row.append(colors[i].name if i < len(colors) else "")
        row.append(all_grades[i] if i < len(all_grades) else "")
        ws_ref.append(row)

    # Template Sheet
    ws_template = wb.create_sheet("Route Template", 0)  # Insert as first sheet
    headers = ["zone_name", "setter_name", "color_name", "intended_grade", "set_date"]
    ws_template.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws_template.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        ws_template.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Example row
    ws_template.append(["Main Wall", "None", "Red", "V4", str(date.today())])

    # --- Add Dropdown Data Validations ---
    # Max allowed rows for bulk import is 500
    validation_rows = 500

    # 1. Zone Name Dropdown (Column A)
    if zones:
        dv_zone = DataValidation(
            type="list",
            formula1=f"Reference!$A$2:$A${len(zones) + 1}",
            allow_blank=False,
        )
        dv_zone.error = "Your entry is not in the list"
        dv_zone.errorTitle = "Invalid Zone"
        ws_template.add_data_validation(dv_zone)
        dv_zone.add(f"A2:A{validation_rows + 1}")

    # 2. Setter Name Dropdown (Column B) - Allow None as an option
    if setters:
        # Add "None" to the reference list manually at the end
        none_row = len(setters) + 2
        ws_ref.cell(row=none_row, column=2, value="None")
        dv_setter = DataValidation(
            type="list", formula1=f"Reference!$B$2:$B${none_row}", allow_blank=True
        )
        dv_setter.error = "Your entry is not in the list"
        dv_setter.errorTitle = "Invalid Setter"
        ws_template.add_data_validation(dv_setter)
        dv_setter.add(f"B2:B{validation_rows + 1}")

    # 3. Color Name Dropdown (Column C)
    if colors:
        dv_color = DataValidation(
            type="list",
            formula1=f"Reference!$C$2:$C${len(colors) + 1}",
            allow_blank=False,
        )
        dv_color.error = "Your entry is not in the list"
        dv_color.errorTitle = "Invalid Color"
        ws_template.add_data_validation(dv_color)
        dv_color.add(f"C2:C{validation_rows + 1}")

    # 4. Intended Grade Dropdown (Column D) - Combine boulder & top rope for simplicity in the master list
    if all_grades:
        dv_grade = DataValidation(
            type="list",
            formula1=f"Reference!$D$2:$D${len(all_grades) + 1}",
            allow_blank=False,
        )
        dv_grade.error = "Your entry is not in the list"
        dv_grade.errorTitle = "Invalid Grade"
        ws_template.add_data_validation(dv_grade)
        dv_grade.add(f"D2:D{validation_rows + 1}")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=route_import_template.xlsx"
        },
    )


@router.post("/routes/bulk-upload", response_model=schemas.BulkUploadResponse)
def bulk_upload_routes(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_super_admin),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    file_bytes = file.file.read()
    if len(file_bytes) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5MB)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    headers = [str(h).strip().lower() for h in rows[0] if h is not None]
    expected_headers = [
        "zone_name",
        "setter_name",
        "color_name",
        "intended_grade",
        "set_date",
    ]

    # Check if expected headers are present
    missing_headers = [h for h in expected_headers if h not in headers]
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing_headers)}",
        )

    # Map column names to indices
    header_indices = {h: i for i, h in enumerate(headers)}

    data_rows = rows[1:]
    if not data_rows:
        raise HTTPException(
            status_code=400, detail="File is empty or contains no data rows"
        )
    if len(data_rows) > 500:
        raise HTTPException(
            status_code=400, detail="Maximum 500 rows allowed per upload"
        )

    errors = []
    routes_to_create = []

    # Prefetch mapping dictionaries for faster validation (case-insensitive)
    zones = {z.name.lower(): z for z in db.query(models.Zone).all()}
    setters = {
        s.name.lower(): s
        for s in db.query(models.Setter).filter(models.Setter.is_active == True).all()
    }
    colors = {c.name.lower(): c for c in db.query(models.Color).all()}

    for i, row in enumerate(data_rows):
        row_num = i + 2  # 1-indexed, plus header

        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue

        def get_val(col_name):
            idx = header_indices.get(col_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        zone_name = get_val("zone_name")
        setter_name = get_val("setter_name")
        color_name = get_val("color_name")
        intended_grade = get_val("intended_grade")
        set_date_raw = get_val("set_date")

        row_has_error = false = False

        # Validate Zone
        zone = None
        if not zone_name:
            errors.append(
                schemas.BulkUploadRowError(
                    row=row_num, field="zone_name", message="Required"
                )
            )
            row_has_error = True
        else:
            zone = zones.get(zone_name.lower())
            if not zone:
                errors.append(
                    schemas.BulkUploadRowError(
                        row=row_num,
                        field="zone_name",
                        message=f"Zone '{zone_name}' not found",
                    )
                )
                row_has_error = True

        # Validate Setter (optional)
        setter_id = None
        if setter_name and setter_name.lower() != "none":
            setter = setters.get(setter_name.lower())
            if not setter:
                errors.append(
                    schemas.BulkUploadRowError(
                        row=row_num,
                        field="setter_name",
                        message=f"Setter '{setter_name}' not found or inactive",
                    )
                )
                row_has_error = True
            else:
                setter_id = setter.id

        # Validate Color
        color_hex = None
        if not color_name:
            errors.append(
                schemas.BulkUploadRowError(
                    row=row_num, field="color_name", message="Required"
                )
            )
            row_has_error = True
        else:
            color = colors.get(color_name.lower())
            if not color:
                errors.append(
                    schemas.BulkUploadRowError(
                        row=row_num,
                        field="color_name",
                        message=f"Color '{color_name}' not found",
                    )
                )
                row_has_error = True
            else:
                color_hex = color.hex_value

        # Validate Grade
        if not intended_grade:
            errors.append(
                schemas.BulkUploadRowError(
                    row=row_num, field="intended_grade", message="Required"
                )
            )
            row_has_error = True
        elif zone:
            valid_grades = GRADES_BY_TYPE.get(zone.route_type, [])
            if intended_grade not in valid_grades:
                errors.append(
                    schemas.BulkUploadRowError(
                        row=row_num,
                        field="intended_grade",
                        message=f"Grade '{intended_grade}' invalid for {zone.route_type} zone",
                    )
                )
                row_has_error = True

        # Parse Set Date
        parsed_date = None
        if set_date_raw:
            try:
                # Value could already be a datetime.datetime object if parsed by openpyxl
                if isinstance(row[header_indices["set_date"]], date):
                    parsed_date = (
                        row[header_indices["set_date"]].date()
                        if hasattr(row[header_indices["set_date"]], "date")
                        else row[header_indices["set_date"]]
                    )
                else:
                    parsed_date = date.fromisoformat(set_date_raw.split("T")[0])
            except ValueError:
                errors.append(
                    schemas.BulkUploadRowError(
                        row=row_num,
                        field="set_date",
                        message="Invalid format. Use YYYY-MM-DD",
                    )
                )
                row_has_error = True

        if not row_has_error:
            # Prepare route object
            db_route = models.Route(
                setter_id=setter_id,
                zone_id=zone.id,
                color=color_hex,
                intended_grade=intended_grade,
                status="active",
            )
            if parsed_date:
                db_route.set_date = parsed_date
            routes_to_create.append(db_route)

    # Atomic creation: only proceed if 0 errors
    if errors:
        return schemas.BulkUploadResponse(
            total_rows=len(data_rows),
            created_count=0,
            error_count=len(errors),
            errors=errors,
        )

    if routes_to_create:
        db.add_all(routes_to_create)
        db.commit()

    return schemas.BulkUploadResponse(
        total_rows=len(data_rows),
        created_count=len(routes_to_create),
        error_count=0,
        errors=[],
    )


# --- Comment Management ---


@router.delete("/comments/{comment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin),
):
    db_comment = (
        db.query(models.Comment).filter(models.Comment.id == comment_id).first()
    )
    if not db_comment:
        raise HTTPException(status_code=404, detail="Comment not found")

    db.delete(db_comment)
    db.commit()
    return None
